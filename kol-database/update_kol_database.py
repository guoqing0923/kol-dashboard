#!/usr/bin/env python3
"""
土耳其KOL达人库更新脚本
用法:
    python3 update_kol_database.py --days 7      # 拉取最近7天数据
    python3 update_kol_database.py --month 2026-05  # 拉取指定月份
    python3 update_kol_database.py --all            # 全量同步TR数据
    python3 update_kol_database.py --init           # 从历史Excel初始化数据库（首次运行）

每次运行（除--init外）自动执行3日/7日定期回查并更新评分。
"""

import argparse
import json
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────
# 路径配置
# ─────────────────────────────────────────
BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / "turkey_kol_database.xlsx"
SECRETS_PATH = Path.home() / ".verdent" / "secrets" / "codeck.json"

HIST_EXCEL = Path.home() / "Downloads" / "土耳其KOL最终评级表-2.xlsx"
APR_EXCEL  = Path.home() / "Downloads" / "4月土耳其达人评级表.xlsx"

API_URL = "https://inner-operations.verdent.ai/inner/talent/promotion/list"

# ─────────────────────────────────────────
# 达人池黑名单（手动删除的达人，防止自动重新入池）
# ─────────────────────────────────────────
POOL_BLACKLIST = {
    "kodumun.muhendisi",
    "cankatko",
    "HQ NET",
    "volkan.js",
    "AI Mevzuları | Yapay Zeka",
}

# ─────────────────────────────────────────
# 评级颜色
# ─────────────────────────────────────────
GRADE_COLORS = {
    "S": "92D050",   # 绿
    "A": "C6EFCE",   # 浅绿
    "B": "FFEB9C",   # 黄
    "C": "FFC7CE",   # 红
}

STAGE_COLORS = {
    "初评":    "DEEBF7",   # 浅蓝
    "3日":     "FFF2CC",   # 浅黄
    "7日终评": "E2EFDA",   # 浅绿
}

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")

# ─────────────────────────────────────────
# 评级标准
# ─────────────────────────────────────────
def score_short_video(vv, cpm, er_pct, platform):
    """短视频评分（满分12分）"""
    # 曝光力
    if vv > 100000:
        s_vv = 4
    elif vv >= 50000:
        s_vv = 3
    elif vv >= 10000:
        s_vv = 2
    else:
        s_vv = 1

    # CPM 得分
    if cpm < 20:
        s_cpm = 4
    elif cpm < 30:
        s_cpm = 3
    elif cpm < 50:
        s_cpm = 2
    else:
        s_cpm = 1

    # ER 得分（TikTok vs IG）
    plat_lower = platform.lower()
    is_tiktok = "tiktok" in plat_lower
    if is_tiktok:
        if er_pct > 10:
            s_er = 4
        elif er_pct >= 6:
            s_er = 3
        elif er_pct >= 3:
            s_er = 2
        else:
            s_er = 1
    else:  # Instagram / 其他
        if er_pct > 4:
            s_er = 4
        elif er_pct >= 2:
            s_er = 3
        elif er_pct >= 0.8:
            s_er = 2
        else:
            s_er = 1

    total = s_vv + s_cpm + s_er
    grade = "S" if total >= 11 else "A" if total >= 8 else "B" if total >= 5 else "C"
    return s_vv, s_cpm, s_er, None, total, 12, grade


def score_long_video(vv, cpm, er_pct, pv):
    """长视频评分（满分16分）"""
    # 曝光力
    if vv > 10000:
        s_vv = 4
    elif vv >= 5000:
        s_vv = 3
    elif vv >= 2000:
        s_vv = 2
    else:
        s_vv = 1

    # CPM 得分
    if cpm < 100:
        s_cpm = 4
    elif cpm < 200:
        s_cpm = 3
    elif cpm < 400:
        s_cpm = 2
    else:
        s_cpm = 1

    # ER 得分
    if er_pct > 7:
        s_er = 4
    elif er_pct >= 5:
        s_er = 3
    elif er_pct >= 2:
        s_er = 2
    else:
        s_er = 1

    # 转化力（PV）
    if pv is not None and str(pv).strip() not in ("", "-", "nan", "NaN"):
        try:
            pv_val = float(pv)
            if pv_val > 100:
                s_pv = 4
            elif pv_val >= 50:
                s_pv = 3
            elif pv_val >= 10:
                s_pv = 2
            else:
                s_pv = 1
        except (ValueError, TypeError):
            s_pv = None
    else:
        s_pv = None

    if s_pv is not None:
        total = s_vv + s_cpm + s_er + s_pv
        max_score = 16
    else:
        total = s_vv + s_cpm + s_er
        max_score = 12

    grade = "S" if total >= 14 else "A" if total >= 10 else "B" if total >= 6 else "C"
    return s_vv, s_cpm, s_er, s_pv, total, max_score, grade


def score_video(video_type, platform, vv, cpm, er_pct, pv=None):
    """统一打分入口"""
    vt = str(video_type).lower()
    is_long = "长" in vt or "youtube" in vt.replace("shorts", "")
    if is_long and "shorts" not in vt:
        return score_long_video(vv, cpm, er_pct, pv)
    else:
        return score_short_video(vv, cpm, er_pct, platform)


def grade_to_suggestion(grade):
    mapping = {
        "S": "必须复投，优先争取独家/季度合作",
        "A": "进入复投候选池，次月优先联系",
        "B": "观察区，可考虑次月再合作一次测试稳定性",
        "C": "原则上淘汰",
    }
    return mapping.get(grade, "")


# ─────────────────────────────────────────
# API 访问
# ─────────────────────────────────────────
def load_token():
    if not SECRETS_PATH.exists():
        print(f"[ERROR] 未找到token文件: {SECRETS_PATH}", file=sys.stderr)
        sys.exit(1)
    with open(SECRETS_PATH) as f:
        return json.load(f)["auth_token"]


def fetch_tr_promotions(token, page=1, page_size=100, filters=None):
    body = {"page": page, "page_size": page_size, "filters": {"country": ["TR"]}}
    if filters:
        body["filters"].update(filters)
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    resp = requests.post(API_URL, headers=headers, json=body, timeout=30)
    resp.raise_for_status()
    return resp.json()["data"]


def is_turkish_kol(item):
    """判断是否为土耳其达人（country=TR 或 language含土耳其语）"""
    country = str(item.get("country", "")).upper()
    language = str(item.get("language", ""))
    return country == "TR" or "土耳其" in language


def fetch_all_tr_promotions(token):
    """分页拉取全部TR达人投放记录（含客户端国家过滤）"""
    first = fetch_tr_promotions(token, page=1, page_size=100)
    total = first["total"]
    items = list(first["list"])
    pages = (total + 99) // 100
    for p in range(2, pages + 1):
        data = fetch_tr_promotions(token, page=p, page_size=100)
        items.extend(data["list"])
    # 客户端二次过滤：确保只保留土耳其达人
    tr_items = [item for item in items if is_turkish_kol(item)]
    return tr_items


def parse_cost(cost_str):
    """解析花费字段（可能含$符号和逗号）"""
    if cost_str is None:
        return 0.0
    s = str(cost_str).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_er(er_str):
    """解析互动率字段（可能含%符号）"""
    if er_str is None:
        return 0.0
    s = str(er_str).replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def compute_initial_stage(publish_date_str):
    """Compute initial review stage based on how old the video is at ingestion time.
    Uses local time and inclusive boundaries:
      >= 7 days  → 7日终评
      >= 3 days  → 3日
      < 3 days   → 初评
    """
    if not publish_date_str or str(publish_date_str).strip() in ("", "nan", "None", "未知"):
        return "初评"
    try:
        pub_date = datetime.strptime(str(publish_date_str)[:10], "%Y-%m-%d")
        days_elapsed = (datetime.now() - pub_date).total_seconds() / 86400
        if days_elapsed >= 7:
            return "7日终评"
        elif days_elapsed >= 3:
            return "3日"
        else:
            return "初评"
    except Exception:
        return "初评"


def api_items_to_df(items, stage="初评"):
    """将API返回的投放记录展开为视频明细DataFrame"""
    now_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M")
    rows = []
    for item in items:
        kol = item["kol"]
        project = item.get("cooperation_project", "")
        cost_total = parse_cost(item.get("cost", 0))
        videos = item.get("videos", [])
        if not videos:
            continue

        # 按视频数均摊花费
        per_video_cost = cost_total / len(videos) if videos else 0

        for v in videos:
            platform = v.get("platform", "")
            publish_date = v.get("publish_date") or ""
            # Bug 1 fix: fall back to create_time date when publish_date is empty
            if not publish_date:
                ct = v.get("create_time") or ""
                if ct:
                    publish_date = ct[:10]  # YYYY-MM-DD
            vv = v.get("exposure_count") or 0
            likes = v.get("like_count") or 0
            comments = v.get("comment_count") or 0
            shares = v.get("share_count") or 0
            collect = v.get("collect_count") or 0

            # 计算互动率
            er_pct = ((likes + comments + shares + collect) / vv * 100) if vv > 0 else 0.0

            # 判断视频类型
            plat_lower = platform.lower()
            if "youtube" in plat_lower and "shorts" not in plat_lower:
                video_type = "长视频"
            else:
                video_type = "短视频"

            # CPM
            cpm = (per_video_cost / vv * 1000) if vv > 0 else 0.0

            # PV（API暂无PV字段）
            pv = None

            # 打分
            s_vv, s_cpm, s_er, s_pv, total, max_score, grade = score_video(
                video_type, platform, vv, cpm, er_pct, pv
            )

            rows.append({
                "达人": kol,
                "宣发期": project,
                "平台": platform,
                "视频类型": video_type,
                "发布日期": publish_date,
                "花费($)": round(per_video_cost, 2),
                "VV曝光量": vv,
                "CPM": round(cpm, 2),
                "互动率(%)": round(er_pct, 2),
                "点赞数": likes,
                "评论数": comments,
                "官网PV": "",
                "曝光力得分": s_vv,
                "CPM得分": s_cpm,
                "ER得分": s_er,
                "PV得分": s_pv if s_pv is not None else "",
                "综合评分": total,
                "满分": max_score,
                "评级": grade,
                "操作建议": grade_to_suggestion(grade),
                "评级阶段": compute_initial_stage(publish_date),
                "最后更新时间": now_str,
                "数据来源": "API",
                "视频ID": v.get("id", ""),
            })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────
# 3日/7日定期回查
# ─────────────────────────────────────────
def refresh_video_data_from_api(old_row, item, v, per_video_cost, now_str):
    """用API最新数据重新计算视频指标和评分，返回更新后的Series副本"""
    new_row = old_row.copy()

    platform = v.get("platform", "") or str(old_row.get("平台", ""))
    vv = v.get("exposure_count") or 0
    likes = v.get("like_count") or 0
    comments = v.get("comment_count") or 0
    shares = v.get("share_count") or 0
    collect = v.get("collect_count") or 0

    er_pct = ((likes + comments + shares + collect) / vv * 100) if vv > 0 else 0.0

    plat_lower = platform.lower()
    if "youtube" in plat_lower and "shorts" not in plat_lower:
        video_type = "长视频"
    else:
        video_type = "短视频"

    cpm = (per_video_cost / vv * 1000) if vv > 0 else 0.0
    pv = None

    s_vv, s_cpm, s_er, s_pv, total, max_score, grade = score_video(
        video_type, platform, vv, cpm, er_pct, pv
    )

    new_row["平台"] = platform
    new_row["视频类型"] = video_type
    new_row["花费($)"] = round(per_video_cost, 2)
    new_row["VV曝光量"] = vv
    new_row["CPM"] = round(cpm, 2)
    new_row["互动率(%)"] = round(er_pct, 2)
    new_row["点赞数"] = likes
    new_row["评论数"] = comments
    new_row["曝光力得分"] = s_vv
    new_row["CPM得分"] = s_cpm
    new_row["ER得分"] = s_er
    new_row["PV得分"] = s_pv if s_pv is not None else ""
    new_row["综合评分"] = total
    new_row["满分"] = max_score
    new_row["评级"] = grade
    new_row["操作建议"] = grade_to_suggestion(grade)
    new_row["最后更新时间"] = now_str
    return new_row


def fix_stale_stages(df):
    """
    Re-evaluate stage for ALL rows where current stage is inconsistent with publish date.
    Handles three correction cases:
      - "7日终评" but video < 7 days old → downgrade to correct stage
      - "初评" but video >= 3 days old   → upgrade to "3日" or "7日终评"
      - "3日" but video < 3 days old     → downgrade to "初评"
    This fixes records wrongly set to "7日终评" by ensure_review_columns (backward compat).
    """
    if "评级阶段" not in df.columns or "发布日期" not in df.columns:
        return df

    now = datetime.now()
    fixed = 0
    for idx in df.index:
        current_stage = str(df.at[idx, "评级阶段"]).strip()
        pub_str = str(df.at[idx, "发布日期"]).strip()
        if not pub_str or pub_str in ("nan", "None", "NaT", "", "未知"):
            continue
        try:
            pub_date = datetime.strptime(pub_str[:10], "%Y-%m-%d")
            days_elapsed = (now - pub_date).total_seconds() / 86400

            if days_elapsed >= 7:
                correct_stage = "7日终评"
            elif days_elapsed >= 3:
                correct_stage = "3日"
            else:
                correct_stage = "初评"

            # Fix if wrong: downgrade 7日终评 that is too young, or upgrade 初评 that is too old
            needs_fix = (
                (current_stage == "7日终评" and days_elapsed < 7) or
                (current_stage == "初评" and days_elapsed >= 3) or
                (current_stage == "3日" and days_elapsed < 3)
            )
            if needs_fix:
                df.at[idx, "评级阶段"] = correct_stage
                fixed += 1
        except Exception:
            pass
    if fixed:
        print(f"  [fix_stages] 修正 {fixed} 条评级阶段与发布日期不一致的记录")
    return df


# Month keyword → zero-padded month number
_MONTH_MAP = {
    "1月": "01", "2月": "02", "3月": "03", "4月": "04",
    "5月": "05", "6月": "06", "7月": "07", "8月": "08",
    "9月": "09", "10月": "10", "11月": "11", "12月": "12",
    "January": "01", "February": "02", "March": "03", "April": "04",
    "May": "05", "June": "06", "July": "07", "August": "08",
    "September": "09", "October": "10", "November": "11", "December": "12",
    "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
    "Jun": "06", "Jul": "07", "Aug": "08", "Sep": "09",
    "Oct": "10", "Nov": "11", "Dec": "12",
}


def infer_date_from_xuanfa(xuanfa_str):
    """Infer a representative date from 宣发期 field, e.g. '4月' → '2026-04-15'."""
    import re
    if not xuanfa_str or str(xuanfa_str).strip() in ("", "nan", "None"):
        return None
    s = str(xuanfa_str).strip()
    # Extract year if present, default to 2026
    year = "2026"
    year_match = re.search(r"(202\d)", s)
    if year_match:
        year = year_match.group(1)
    # Match month keyword (longest first to avoid partial matches)
    for kw in sorted(_MONTH_MAP.keys(), key=len, reverse=True):
        if kw in s:
            return f"{year}-{_MONTH_MAP[kw]}-15"
    return None


def fill_missing_publish_dates(df, api_items):
    """
    Fill empty 发布日期 in df using:
      1. API lookup by (kol_name, platform) exact match → use latest date found
      2. API lookup by kol_name only with partial platform match
      3. Infer from 宣发期 field (e.g. '4月' → '2026-04-15')
      4. Fall back to '未知'
    Only modifies rows where 发布日期 is blank/nan.
    """
    if "发布日期" not in df.columns:
        return df

    # Build API lookup tables
    api_kol_plat = {}   # (kol_lower, plat_lower) → [date_str, ...]
    api_kol_only = {}   # kol_lower → [(plat_lower, date_str), ...]

    for item in api_items:
        kol = str(item.get("kol", "")).strip().lower()
        if not kol:
            continue
        for v in item.get("videos", []):
            plat = str(v.get("platform", "")).strip().lower()
            pub = v.get("publish_date") or ""
            if not pub:
                ct = v.get("create_time") or ""
                if ct:
                    pub = ct[:10]
            if pub and len(pub) >= 10:
                pub = pub[:10]
                key = (kol, plat)
                api_kol_plat.setdefault(key, []).append(pub)
                api_kol_only.setdefault(kol, []).append((plat, pub))

    filled_api = 0
    filled_infer = 0
    filled_unknown = 0

    for idx in df.index:
        pub_val = str(df.at[idx, "发布日期"]).strip()
        if pub_val and pub_val not in ("", "nan", "None", "NaT"):
            continue  # already has a date

        kol = str(df.at[idx, "达人"]).strip().lower()
        plat = str(df.at[idx, "平台"]).strip().lower()
        xuanfa = str(df.at[idx, "宣发期"]).strip()

        # 1. Exact (kol, platform) match
        key = (kol, plat)
        if key in api_kol_plat and api_kol_plat[key]:
            best = sorted(api_kol_plat[key])[-1]
            df.at[idx, "发布日期"] = best
            filled_api += 1
            continue

        # 2. kol-only match with best platform similarity
        if kol in api_kol_only:
            candidates = api_kol_only[kol]
            # Prefer platform match (first 4 chars)
            plat_pref = [d for p, d in candidates if plat[:4] in p or p[:4] in plat]
            pool = plat_pref if plat_pref else [d for _, d in candidates]
            if pool:
                best = sorted(pool)[-1]
                df.at[idx, "发布日期"] = best
                filled_api += 1
                continue

        # 3. Infer from 宣发期
        inferred = infer_date_from_xuanfa(xuanfa)
        if inferred:
            df.at[idx, "发布日期"] = inferred
            filled_infer += 1
            continue

        # 4. Unknown
        df.at[idx, "发布日期"] = "未知"
        filled_unknown += 1

    total = filled_api + filled_infer + filled_unknown
    if total:
        print(f"  [fill_dates] 共补填 {total} 条空发布日期: "
              f"API匹配={filled_api} | 宣发期推算={filled_infer} | 标注未知={filled_unknown}")
    return df


def run_periodic_reviews(df, all_items):
    """
    执行3日和7日定期回查，更新视频数据和评级阶段。
    返回 (updated_df, updated_3day_list, updated_7day_list)
    """
    now = datetime.utcnow()
    now_str = now.strftime("%Y-%m-%d %H:%M")

    # 向后兼容：补全新字段
    if "评级阶段" not in df.columns:
        df = df.copy()
        df["评级阶段"] = "7日终评"
    if "最后更新时间" not in df.columns:
        df = df.copy()
        df["最后更新时间"] = ""

    # 构建 API 视频查找表：video_id -> (item, v, per_video_cost)
    api_video_lookup = {}
    for item in all_items:
        cost_total = parse_cost(item.get("cost", 0))
        videos = item.get("videos", [])
        per_video_cost = cost_total / len(videos) if videos else 0
        for v in videos:
            vid = str(v.get("id", "")).strip()
            if vid and vid not in ("", "0", "nan"):
                api_video_lookup[vid] = (item, v, per_video_cost)

    updated_3day = []
    updated_7day = []

    for idx in df.index:
        row = df.loc[idx]
        stage = str(row.get("评级阶段", "")).strip()
        pub_date_str = str(row.get("发布日期", "")).strip()

        if not pub_date_str or pub_date_str in ("nan", "", "NaT", "None"):
            continue

        # 解析发布日期（只取前10位 YYYY-MM-DD）
        try:
            pub_date = datetime.strptime(pub_date_str[:10], "%Y-%m-%d")
        except Exception:
            continue

        days_elapsed = (now - pub_date).total_seconds() / 86400

        # ── 3日回查：发布3天 ±12h，评级阶段为「初评」 ──
        if stage == "初评" and 2.5 <= days_elapsed <= 3.5:
            vid = str(row.get("视频ID", "")).strip()
            old_grade = str(row.get("评级", ""))

            if vid in api_video_lookup:
                item_api, v_api, pvc = api_video_lookup[vid]
                new_row = refresh_video_data_from_api(row, item_api, v_api, pvc, now_str)
                new_row["评级阶段"] = "3日"
                df.loc[idx] = new_row
                new_grade = str(new_row["评级"])
                api_updated = True
            else:
                df.at[idx, "评级阶段"] = "3日"
                df.at[idx, "最后更新时间"] = now_str
                new_grade = old_grade
                api_updated = False

            updated_3day.append({
                "达人": str(row.get("达人", "")),
                "视频ID": vid,
                "旧评级": old_grade,
                "新评级": new_grade,
                "API更新": api_updated,
            })

        # ── 7日终评：发布7天 ±12h，评级阶段为「3日」 ──
        elif stage == "3日" and 6.5 <= days_elapsed <= 7.5:
            vid = str(row.get("视频ID", "")).strip()
            old_grade = str(row.get("评级", ""))

            if vid in api_video_lookup:
                item_api, v_api, pvc = api_video_lookup[vid]
                new_row = refresh_video_data_from_api(row, item_api, v_api, pvc, now_str)
                new_row["评级阶段"] = "7日终评"
                df.loc[idx] = new_row
                new_grade = str(new_row["评级"])
                api_updated = True
            else:
                df.at[idx, "评级阶段"] = "7日终评"
                df.at[idx, "最后更新时间"] = now_str
                new_grade = old_grade
                api_updated = False

            updated_7day.append({
                "达人": str(row.get("达人", "")),
                "视频ID": vid,
                "旧评级": old_grade,
                "新评级": new_grade,
                "API更新": api_updated,
            })

    return df, updated_3day, updated_7day


# ─────────────────────────────────────────
# 历史Excel 解析
# ─────────────────────────────────────────
def load_history_detail():
    """读取历史Excel每期明细"""
    now_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M")
    df = pd.read_excel(HIST_EXCEL, sheet_name="每期明细")
    df = df.dropna(subset=["达人"])
    rows = []
    for _, r in df.iterrows():
        vv = r.get("曝光量", 0) or 0
        cpm = r.get("CPM", 0) or 0
        er_pct = r.get("互动率(%)", 0) or 0
        if isinstance(er_pct, str):
            er_pct = parse_er(er_pct)
        pv = r.get("官网PV", None)
        platform = str(r.get("主平台", ""))
        video_type = str(r.get("视频类型", ""))
        likes = r.get("点赞数", 0) or 0
        comments = r.get("评论数", 0) or 0

        # 优先用原始得分，如缺失则重新计算
        s_vv = r.get("曝光力", None)
        s_cpm = r.get("CPM得分", None)
        s_er = r.get("互动力", None)
        s_pv_raw = r.get("转化力", None)
        s_pv = None if (s_pv_raw in [None, "-", ""]) or (isinstance(s_pv_raw, float) and pd.isna(s_pv_raw)) else s_pv_raw
        total = r.get("综合评分", None)
        grade = r.get("最终等级", None)

        if None in (s_vv, s_cpm, s_er, total, grade) or str(grade).strip() == "nan":
            s_vv, s_cpm, s_er, s_pv, total, max_score, grade = score_video(
                video_type, platform, vv, cpm, er_pct, pv
            )
        else:
            # 判断满分
            vt = str(video_type).lower()
            is_long = "长" in vt and "shorts" not in vt
            max_score = 16 if is_long else 12

        rows.append({
            "达人": r.get("达人", ""),
            "宣发期": r.get("宣发期", ""),
            "平台": platform,
            "视频类型": video_type,
            "发布日期": "",
            "花费($)": r.get("花费($)", 0),
            "VV曝光量": vv,
            "CPM": round(float(cpm), 2) if cpm else 0,
            "互动率(%)": round(float(er_pct), 2),
            "点赞数": likes,
            "评论数": comments,
            "官网PV": pv if pv not in [None, "-"] else "",
            "曝光力得分": s_vv,
            "CPM得分": s_cpm,
            "ER得分": s_er,
            "PV得分": s_pv if s_pv is not None else "",
            "综合评分": total,
            "满分": max_score,
            "评级": grade,
            "操作建议": grade_to_suggestion(str(grade)),
            "评级阶段": "7日终评",
            "最后更新时间": now_str,
            "数据来源": "历史",
            "视频ID": "",
        })
    return pd.DataFrame(rows)


def load_april_detail():
    """读取4月Excel视频明细"""
    now_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M")
    try:
        df = pd.read_excel(APR_EXCEL, sheet_name="视频明细")
    except Exception:
        return pd.DataFrame()

    df = df.dropna(subset=["KOL名称"])
    rows = []
    for _, r in df.iterrows():
        kol = r.get("KOL名称", "")
        project = r.get("合作项目", "")
        platform = str(r.get("平台", ""))
        video_type = str(r.get("视频类型", ""))
        # 标准化视频类型
        if "长视频" in video_type or "YouTube" in video_type:
            video_type = "长视频"
        else:
            video_type = "短视频"

        publish_date = r.get("发布日期", "")
        if pd.notna(publish_date) and hasattr(publish_date, "strftime"):
            publish_date = publish_date.strftime("%Y-%m-%d")

        cost_raw = r.get("花费(均摊)", 0)
        cost = parse_cost(cost_raw)

        vv = r.get("VV曝光量", 0) or 0
        cpm_raw = r.get("CPM", 0)
        cpm = parse_cost(cpm_raw)

        er_raw = r.get("互动率ER", 0)
        er_pct = parse_er(er_raw) if er_raw else 0.0

        pv = r.get("官网PV", None)

        # 优先用表格里的得分
        s_vv = r.get("曝光力得分", None)
        s_cpm = r.get("CPM得分", None)
        s_er = r.get("ER得分", None)
        s_pv_raw = r.get("PV得分", None)
        s_pv = None if (s_pv_raw in [None, "", "-"] or (isinstance(s_pv_raw, float) and pd.isna(s_pv_raw))) else s_pv_raw
        total = r.get("总分", None)
        max_score = r.get("满分", None)
        grade = r.get("评级", None)

        if None in (s_vv, s_cpm, s_er, total, grade) or str(grade).strip() == "nan":
            s_vv, s_cpm, s_er, s_pv, total, max_score, grade = score_video(
                video_type, platform, vv, cpm, er_pct, pv
            )

        suggestion = r.get("操作建议", "") or grade_to_suggestion(str(grade))

        rows.append({
            "达人": kol,
            "宣发期": project,
            "平台": platform,
            "视频类型": video_type,
            "发布日期": str(publish_date) if publish_date and str(publish_date) != "nan" else "",
            "花费($)": round(cost, 2),
            "VV曝光量": vv,
            "CPM": round(cpm, 2),
            "互动率(%)": round(er_pct, 2),
            "点赞数": 0,
            "评论数": 0,
            "官网PV": pv if pv not in [None, "-", "NaN"] and not (isinstance(pv, float) and pd.isna(pv)) else "",
            "曝光力得分": s_vv,
            "CPM得分": s_cpm,
            "ER得分": s_er,
            "PV得分": s_pv if s_pv is not None else "",
            "综合评分": total,
            "满分": max_score,
            "评级": grade,
            "操作建议": suggestion,
            "评级阶段": "7日终评",
            "最后更新时间": now_str,
            "数据来源": "4月",
            "视频ID": "",
        })
    return pd.DataFrame(rows)


def load_pool():
    """读取历史达人池"""
    df = pd.read_excel(HIST_EXCEL, sheet_name="达人池")
    df = df.dropna(subset=["达人"])
    return df


# ─────────────────────────────────────────
# 达人池自动更新
# ─────────────────────────────────────────
_GRADE_ORDER = {"S": 0, "A": 1, "B": 2, "C": 3}


def _best_grade(grades):
    """返回一组评级中最高的（S > A > B > C）"""
    valid = [g for g in grades if g in _GRADE_ORDER]
    if not valid:
        return "C"
    return min(valid, key=lambda g: _GRADE_ORDER[g])


def _infer_video_type(has_short, has_long):
    if has_short and has_long:
        return "混合"
    elif has_long:
        return "长视频"
    else:
        return "短视频"


def _collect_platforms(platform_series):
    """从平台字段（可能是逗号分隔的多平台字符串）收集所有去重后的平台"""
    seen = []
    for raw in platform_series.dropna():
        for part in str(raw).split(","):
            p = part.strip()
            if p and p not in seen:
                seen.append(p)
    return seen


def update_pool_from_detail(detail_df, pool_df):
    """
    根据视频明细自动更新达人池：
    - 短视频(TikTok/Instagram/Shorts) S级 → 入池
    - 长视频(YouTube) B级及以上(S/A/B) → 入池
    - 已在池中的达人：保留手工字段，更新评级/视频类型/发布平台/过往单次花费
    - 不再符合条件的达人：保留，在备注中标注（不删除）
    返回更新后的 pool_df
    """
    if detail_df.empty:
        return pool_df

    # ── 手工保留字段（不自动覆盖）──
    MANUAL_FIELDS = ["粉丝量", "联系方式"]
    AUTO_FIELDS   = ["视频类型", "发布平台", "过往单次花费", "评级"]
    POOL_COLS     = ["达人", "视频类型", "发布平台", "过往单次花费", "评级",
                     "粉丝量", "联系方式"]

    # ── 构建现有池的手工信息字典：kol_name → {field: value} ──
    existing_manual = {}
    existing_in_pool = set()
    if not pool_df.empty:
        for _, row in pool_df.iterrows():
            kol = str(row.get("达人", "")).strip()
            if not kol:
                continue
            existing_in_pool.add(kol)
            existing_manual[kol] = {
                f: (str(row.get(f, "")) if pd.notna(row.get(f, "")) else "")
                for f in MANUAL_FIELDS
            }

    # ── 扫描视频明细，找出符合条件的达人 ──
    qualifying_kols = set()
    for kol, grp in detail_df.groupby("达人"):
        kol_name = str(kol).strip()
        # 黑名单达人不得入池
        if kol_name in POOL_BLACKLIST:
            continue
        short_grades = grp.loc[grp["视频类型"] == "短视频", "评级"].tolist()
        long_grades  = grp.loc[grp["视频类型"] == "长视频",  "评级"].tolist()
        ok_short = any(g == "S" for g in short_grades)
        ok_long  = any(g in ("S", "A", "B") for g in long_grades)
        if ok_short or ok_long:
            qualifying_kols.add(kol_name)

    print(f"  [达人池] 符合条件达人: {len(qualifying_kols)} 人 | "
          f"已在池中: {len(existing_in_pool)} 人")

    # ── 为每个符合条件的达人汇总数据 ──
    new_pool_rows = []
    for kol in sorted(qualifying_kols):
        grp = detail_df[detail_df["达人"] == kol]
        short_grp = grp[grp["视频类型"] == "短视频"]
        long_grp  = grp[grp["视频类型"] == "长视频"]

        has_short = not short_grp.empty
        has_long  = not long_grp.empty
        video_type = _infer_video_type(has_short, has_long)

        all_platforms = _collect_platforms(grp["平台"])
        platform_str  = ", ".join(all_platforms)

        # 过往单次花费：取平均（保留2位小数）
        costs = grp["花费($)"].dropna()
        avg_cost = round(float(costs.mean()), 2) if not costs.empty else 0.0

        # 评级：取最高等级
        all_grades = grp["评级"].dropna().tolist()
        best = _best_grade(all_grades)

        # 手工字段：从现有池继承，新达人留空
        manual = existing_manual.get(kol, {f: "" for f in MANUAL_FIELDS})

        row = {"达人": kol, "视频类型": video_type, "发布平台": platform_str,
               "过往单次花费": avg_cost, "评级": best}
        row.update(manual)
        new_pool_rows.append(row)

    # ── 处理已在池中但不再符合条件的达人（保留）──
    no_longer_qualifying = existing_in_pool - qualifying_kols
    if no_longer_qualifying:
        print(f"  [达人池] 不再符合条件（保留）: {sorted(no_longer_qualifying)}")
    for kol in no_longer_qualifying:
        if not pool_df.empty:
            old_rows = pool_df[pool_df["达人"] == kol]
            for _, old_row in old_rows.iterrows():
                row = {
                    c: (str(old_row.get(c, "")) if pd.notna(old_row.get(c, "")) else "")
                    for c in POOL_COLS
                }
                new_pool_rows.append(row)

    # ── 构建新 pool_df，确保列顺序一致 ──
    new_pool_df = pd.DataFrame(new_pool_rows, columns=POOL_COLS)

    # 按评级排序（S最前）
    new_pool_df["_sort"] = new_pool_df["评级"].map(_GRADE_ORDER).fillna(9)
    new_pool_df = new_pool_df.sort_values("_sort").drop(columns="_sort").reset_index(drop=True)

    added   = len(qualifying_kols - existing_in_pool)
    updated = len(qualifying_kols & existing_in_pool)
    print(f"  [达人池] 新增 {added} 人 | 更新 {updated} 人 | 标注不符合 {len(no_longer_qualifying)} 人 | "
          f"池总人数: {len(new_pool_df)}")
    return new_pool_df


# ─────────────────────────────────────────
# 达人信息自动搜集（粉丝量 / 联系方式）
# ─────────────────────────────────────────
def _format_followers(count):
    """格式化粉丝数为易读形式"""
    try:
        n = int(count)
        if n >= 1_000_000:
            return f"{n / 1_000_000:.1f}M"
        elif n >= 1_000:
            return f"{n / 1_000:.1f}K"
        return str(n)
    except (ValueError, TypeError):
        return str(count)


def _search_kol_info(kol_name):
    """
    尝试从互联网搜集KOL的粉丝量和联系方式。
    依次尝试：Instagram页面 → SocialBlade。
    返回 (fans_str, contact_str)，找不到则返回 ("", "")。
    每次调用最多发出 2 个 HTTP 请求，超时 10 秒。
    """
    import re

    fans = ""
    contact = ""

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }

    # 1. 尝试 Instagram 页面
    try:
        url = f"https://www.instagram.com/{kol_name}/"
        resp = requests.get(url, headers=headers, timeout=10, allow_redirects=True)
        if resp.status_code == 200:
            text = resp.text
            # 粉丝数（JSON 片段）
            m = re.search(r'"edge_followed_by"\s*:\s*\{"count"\s*:\s*(\d+)', text)
            if not m:
                # 备用：meta description 里的数字
                m = re.search(r'([\d,]+)\s+Followers', text)
            if m:
                raw = m.group(1).replace(",", "")
                fans = _format_followers(raw)
            # 邮箱（Instagram bio 里可能有）
            emails = re.findall(
                r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', text
            )
            _skip = {"sentry", "instagram", "facebook", "example", "test",
                     "support", "privacy", "noreply", "help"}
            for email in emails:
                if not any(s in email.lower() for s in _skip):
                    contact = email
                    break
    except Exception:
        pass

    # 2. 如果 Instagram 没拿到粉丝量，尝试 SocialBlade
    if not fans:
        try:
            url = f"https://socialblade.com/instagram/user/{kol_name}"
            resp = requests.get(url, headers=headers, timeout=10, allow_redirects=True)
            if resp.status_code == 200:
                text = resp.text
                m = re.search(r'Followers[^<]{0,60}?([0-9][0-9,\.]+[KMBkmb]?)', text)
                if m:
                    fans = m.group(1).strip()
        except Exception:
            pass

    return fans, contact


def collect_kol_info(pool_df):
    """
    对粉丝量或联系方式为空的达人，自动从互联网搜集信息。
    已有完整数据的行不重复搜集。
    搜集失败不阻塞流程。
    """
    if pool_df.empty:
        return pool_df

    need_collect = []
    for idx in pool_df.index:
        kol = str(pool_df.at[idx, "达人"]).strip()
        if not kol or kol in ("nan", "None", ""):
            continue
        fans = str(pool_df.at[idx, "粉丝量"]).strip() if "粉丝量" in pool_df.columns else ""
        contact = str(pool_df.at[idx, "联系方式"]).strip() if "联系方式" in pool_df.columns else ""
        fans_empty = fans in ("", "nan", "None", "0")
        contact_empty = contact in ("", "nan", "None")
        if fans_empty or contact_empty:
            need_collect.append((idx, kol, fans_empty, contact_empty))

    if not need_collect:
        return pool_df

    print(f"  [搜集] 需要补充信息的达人: {len(need_collect)} 人")
    updated = 0
    for idx, kol, fans_empty, contact_empty in need_collect:
        print(f"  [搜集] {kol} ...", end=" ", flush=True)
        try:
            new_fans, new_contact = _search_kol_info(kol)
            changed = []
            if fans_empty and new_fans:
                pool_df.at[idx, "粉丝量"] = new_fans
                changed.append(f"粉丝量={new_fans}")
            if contact_empty and new_contact:
                pool_df.at[idx, "联系方式"] = new_contact
                changed.append(f"联系方式={new_contact}")
            if changed:
                updated += 1
                print(" | ".join(changed))
            else:
                print("未找到")
        except Exception as e:
            print(f"失败: {e}")

    print(f"  [搜集] 共补充 {updated} 位达人的信息")
    return pool_df


# ─────────────────────────────────────────
# 达人汇总计算
# ─────────────────────────────────────────
def compute_summary(detail_df):
    """按达人聚合视频明细，生成达人汇总"""
    if detail_df.empty:
        return pd.DataFrame()

    rows = []
    for kol, grp in detail_df.groupby("达人"):
        cost = grp["花费($)"].sum()
        total_vv = grp["VV曝光量"].sum()
        avg_cpm = (cost / total_vv * 1000) if total_vv > 0 else 0
        # 最新等级 = 最后一条记录的评级
        grp_sorted = grp.sort_values("发布日期", na_position="first")
        latest_grade = grp_sorted.iloc[-1]["评级"]
        # 等级趋势
        grades = grp_sorted["评级"].tolist()
        if len(grades) > 1:
            trend_str = " → ".join(grades)
            if grades[-1] > grades[0]:
                trend = "UP"
            elif grades[-1] < grades[0]:
                trend = "DOWN"
            else:
                trend = "STABLE"
        else:
            trend_str = grades[0] if grades else ""
            trend = "-"

        # 操作建议（最新等级为准）
        suggestion = grade_to_suggestion(str(latest_grade))

        # 主平台
        platforms = grp["平台"].dropna().tolist()
        main_platform = max(set(platforms), key=platforms.count) if platforms else ""

        # 最新评级阶段
        latest_stage = grp_sorted.iloc[-1].get("评级阶段", "") if "评级阶段" in grp_sorted.columns else ""

        rows.append({
            "达人": kol,
            "主平台": main_platform,
            "合作次数": len(grp),
            "宣发期列表": " | ".join(grp["宣发期"].dropna().unique().tolist()),
            "总花费($)": round(cost, 2),
            "总VV曝光量": total_vv,
            "平均CPM": round(avg_cpm, 2),
            "各期评级": trend_str,
            "等级趋势": trend,
            "最新等级": latest_grade,
            "最新评级阶段": latest_stage,
            "操作建议": suggestion,
        })

    df = pd.DataFrame(rows)
    grade_order = {"S": 0, "A": 1, "B": 2, "C": 3}
    df["_sort"] = df["最新等级"].map(grade_order).fillna(9)
    df = df.sort_values("_sort").drop(columns="_sort").reset_index(drop=True)
    return df


# ─────────────────────────────────────────
# Excel 写入工具
# ─────────────────────────────────────────
def _thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_table_style(ws, df, grade_col=None, stage_col=None, freeze_row=2):
    """为worksheet应用表格样式"""
    # 冻结首行
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)

    # 表头样式
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()

    # 数据行
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for cell in row:
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if cell.fill.fgColor.rgb == "00000000":
                cell.fill = fill

    # 评级列着色
    header_vals = [c.value for c in ws[1]]
    if grade_col and grade_col in header_vals:
        col_idx = header_vals.index(grade_col) + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cell = row[col_idx - 1]
            grade = str(cell.value).strip() if cell.value else ""
            if grade in GRADE_COLORS:
                cell.fill = PatternFill("solid", fgColor=GRADE_COLORS[grade])
                cell.font = Font(bold=True)

    # 评级阶段列着色
    if stage_col and stage_col in header_vals:
        col_idx = header_vals.index(stage_col) + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cell = row[col_idx - 1]
            stage = str(cell.value).strip() if cell.value else ""
            if stage in STAGE_COLORS:
                cell.fill = PatternFill("solid", fgColor=STAGE_COLORS[stage])

    # 自适应列宽（最大40）
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value else 0) for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 8), 40)


def write_excel(detail_df, summary_df, pool_df, path):
    """写入主数据库Excel"""
    wb = Workbook()

    # Sheet 1: 视频明细
    ws_detail = wb.active
    ws_detail.title = "视频明细"
    if not detail_df.empty:
        cols = list(detail_df.columns)
        ws_detail.append(cols)
        for _, row in detail_df.iterrows():
            ws_detail.append([row.get(c, "") for c in cols])
        apply_table_style(ws_detail, detail_df, grade_col="评级", stage_col="评级阶段")
    ws_detail.row_dimensions[1].height = 30

    # Sheet 2: 达人汇总
    ws_summary = wb.create_sheet("达人汇总")
    if not summary_df.empty:
        cols = list(summary_df.columns)
        ws_summary.append(cols)
        for _, row in summary_df.iterrows():
            ws_summary.append([row.get(c, "") for c in cols])
        apply_table_style(ws_summary, summary_df, grade_col="最新等级", stage_col="最新评级阶段")
    ws_summary.row_dimensions[1].height = 30

    # Sheet 3: 达人池
    ws_pool = wb.create_sheet("达人池")
    if not pool_df.empty:
        cols = list(pool_df.columns)
        ws_pool.append(cols)
        for _, row in pool_df.iterrows():
            ws_pool.append([row.get(c, "") for c in cols])
        apply_table_style(ws_pool, pool_df, grade_col="评级")
    ws_pool.row_dimensions[1].height = 30

    wb.save(path)
    print(f"[✓] 数据库已保存: {path}")
    # 同时生成HTML看板
    dashboard_path = Path(path).parent / "dashboard.html"
    generate_html_dashboard(detail_df, summary_df, pool_df, dashboard_path)


# ─────────────────────────────────────────
# HTML看板生成
# ─────────────────────────────────────────
def generate_html_dashboard(detail_df, summary_df, pool_df, output_path):
    """生成单文件HTML看板，数据嵌入JSON，支持排序/搜索/评级着色"""
    import json as _json

    def _safe(v):
        if v is None:
            return ""
        if isinstance(v, float):
            if pd.isna(v):
                return ""
            return int(v) if v == int(v) else round(float(v), 2)
        if isinstance(v, int):
            return v
        s = str(v)
        return "" if s in ("nan", "NaT", "None", "NaN") else s

    def _to_records(df):
        if df is None or df.empty:
            return [], []
        cols = list(df.columns)
        records = [{c: _safe(row[c]) for c in cols} for _, row in df.iterrows()]
        return cols, records

    dc, dd = _to_records(detail_df)
    sc, sd = _to_records(summary_df)
    pc, pool_records = _to_records(pool_df)

    total_kols   = len(sd)
    total_videos = len(dd)
    try:
        total_cost = sum(float(r.get("花费($)", 0) or 0) for r in dd)
        total_vv   = sum(int(r.get("VV曝光量", 0) or 0) for r in dd)
    except Exception:
        total_cost, total_vv = 0, 0
    avg_cpm = round(total_cost / total_vv * 1000, 2) if total_vv > 0 else 0

    gd = {"S": 0, "A": 0, "B": 0, "C": 0}
    for r in sd:
        g = str(r.get("最新等级", "")).strip()
        if g in gd:
            gd[g] += 1

    # 评级阶段统计
    stage_counts = {"初评": 0, "3日": 0, "7日终评": 0}
    for r in dd:
        s = str(r.get("评级阶段", "")).strip()
        if s in stage_counts:
            stage_counts[s] += 1

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    tabs_json = _json.dumps([
        {"id": "detail",  "label": "视频明细", "cols": dc, "data": dd,           "gradeKey": "评级"},
        {"id": "summary", "label": "达人汇总", "cols": sc, "data": sd,           "gradeKey": "最新等级"},
        {"id": "pool",    "label": "达人池",   "cols": pc, "data": pool_records, "gradeKey": "评级"},
    ], ensure_ascii=False)

    stats_json = _json.dumps([
        {"label": "总达人数",  "value": str(total_kols)},
        {"label": "总视频数",  "value": str(total_videos)},
        {"label": "总花费($)", "value": "${:,.0f}".format(total_cost)},
        {"label": "平均CPM",  "value": "${:.2f}".format(avg_cpm)},
    ], ensure_ascii=False)

    data_block = (
        "const TABS=" + tabs_json + ";\n"
        "const STATS_CARDS=" + stats_json + ";\n"
        "const GRADE_DIST=" + _json.dumps(gd) + ";\n"
        "const STAGE_COUNTS=" + _json.dumps(stage_counts) + ";\n"
        "const UPDATE_TIME=" + _json.dumps(now_str) + ";\n"
    )
    # Prevent </script> injection
    data_block = data_block.replace("</script>", "<\\/script>")

    css = (
        "*{box-sizing:border-box;margin:0;padding:0}"
        "body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f6fa;color:#333}"
        "header{background:linear-gradient(135deg,#1E3A5F,#2980b9);color:#fff;padding:20px 24px}"
        "header h1{font-size:1.4rem;font-weight:700}"
        "header p{font-size:.8rem;opacity:.75;margin-top:4px}"
        ".stats{display:flex;flex-wrap:wrap;gap:12px;padding:20px 24px 8px}"
        ".stat-card{background:#fff;border-radius:10px;padding:14px 18px;min-width:130px;flex:1;"
        "box-shadow:0 2px 8px rgba(0,0,0,.08)}"
        ".stat-card .lbl{font-size:.72rem;color:#888;text-transform:uppercase;letter-spacing:.05em}"
        ".stat-card .val{font-size:1.5rem;font-weight:700;color:#1E3A5F;margin-top:4px}"
        ".grade-row{display:flex;flex-wrap:wrap;gap:8px;padding:8px 24px 4px}"
        ".stage-row{display:flex;flex-wrap:wrap;gap:8px;padding:4px 24px 16px}"
        ".gbadge{padding:5px 14px;border-radius:20px;font-weight:700;font-size:.85rem;color:#fff}"
        ".sbadge{padding:4px 12px;border-radius:20px;font-size:.8rem;font-weight:600}"
        ".gs{background:#2E7D32}.ga{background:#66BB6A}.gb{background:#FFA726}.gc2{background:#EF5350}"
        ".ss-init{background:#DEEBF7;color:#1565C0}"
        ".ss-3d{background:#FFF2CC;color:#E65100}"
        ".ss-7d{background:#E2EFDA;color:#2E7D32}"
        ".tabs{display:flex;background:#fff;border-bottom:2px solid #dde1e7;padding:0 24px}"
        ".tbtn{padding:12px 22px;cursor:pointer;border:none;background:none;font-size:.9rem;"
        "color:#666;border-bottom:3px solid transparent;margin-bottom:-2px;transition:.2s}"
        ".tbtn.active{color:#1E3A5F;border-bottom-color:#2980b9;font-weight:600}"
        ".tbtn:hover{background:#f0f4f8;color:#1E3A5F}"
        ".panel{display:none;padding:18px 24px}.panel.active{display:block}"
        ".toolbar{display:flex;align-items:center;gap:10px;margin-bottom:12px;flex-wrap:wrap}"
        ".sinput{padding:7px 12px;border:1px solid #ddd;border-radius:6px;font-size:.88rem;"
        "width:260px;outline:none}"
        ".sinput:focus{border-color:#2980b9;box-shadow:0 0 0 2px rgba(41,128,185,.2)}"
        ".rc{font-size:.82rem;color:#888}"
        ".tw{overflow-x:auto;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,.08)}"
        "table{width:100%;border-collapse:collapse;background:#fff;font-size:.82rem}"
        "thead th{background:#1E3A5F;color:#fff;padding:9px 11px;text-align:left;cursor:pointer;"
        "user-select:none;white-space:nowrap;position:sticky;top:0;z-index:1}"
        "thead th:hover{background:#2c5282}"
        ".si{margin-left:3px;font-size:.7rem;opacity:.7}"
        ".sa .si::after{content:'▲'}.sd .si::after{content:'▼'}"
        "th:not(.sa):not(.sd) .si::after{content:'⇅';opacity:.4}"
        "tbody tr:nth-child(even){background:#f8f9ff}"
        "tbody tr:hover{background:#e8f0fe!important}"
        "tbody td{padding:7px 11px;border-bottom:1px solid #eee;white-space:nowrap;"
        "max-width:220px;overflow:hidden;text-overflow:ellipsis}"
        ".badge{display:inline-block;padding:2px 9px;border-radius:10px;font-size:.78rem;font-weight:700}"
        ".sbadge-cell{display:inline-block;padding:2px 8px;border-radius:8px;font-size:.76rem;font-weight:600}"
        ".gc-S{background:#E8F5E9!important;color:#2E7D32}.badge.grade-S{background:#2E7D32;color:#fff}"
        ".gc-A{background:#F1F8E9!important;color:#388E3C}.badge.grade-A{background:#66BB6A;color:#fff}"
        ".gc-B{background:#FFF8E1!important;color:#E65100}.badge.grade-B{background:#FFA726;color:#fff}"
        ".gc-C{background:#FFEBEE!important;color:#C62828}.badge.grade-C{background:#EF5350;color:#fff}"
        ".stage-初评{background:#DEEBF7;color:#1565C0}"
        ".stage-3日{background:#FFF2CC;color:#E65100}"
        ".stage-7日终评{background:#E2EFDA;color:#2E7D32}"
        "@media(max-width:768px){.stats,.panel{padding:12px}.sinput{width:100%}header{padding:14px 16px}}"
    )

    js = r"""
document.getElementById("utime").textContent = "数据更新时间: " + UPDATE_TIME;

// Stats cards
var sa = document.getElementById("stats-area");
STATS_CARDS.forEach(function(s){
  var d = document.createElement("div"); d.className = "stat-card";
  d.innerHTML = '<div class="lbl">'+s.label+'</div><div class="val">'+s.value+'</div>';
  sa.appendChild(d);
});

// Grade badges
var ga = document.getElementById("grade-area");
var GRADE_CSS = {S:"gs",A:"ga",B:"gb",C:"gc2"};
["S","A","B","C"].forEach(function(g){
  var span = document.createElement("span");
  span.className = "gbadge " + GRADE_CSS[g];
  span.textContent = g + " · " + GRADE_DIST[g] + " 人";
  ga.appendChild(span);
});

// Stage badges
var stageArea = document.getElementById("stage-area");
var STAGE_LABELS = [
  {key:"初评",   css:"ss-init", label:"初评"},
  {key:"3日",    css:"ss-3d",   label:"3日回查"},
  {key:"7日终评",css:"ss-7d",   label:"7日终评"},
];
STAGE_LABELS.forEach(function(st){
  var span = document.createElement("span");
  span.className = "sbadge " + st.css;
  span.textContent = st.label + " · " + (STAGE_COUNTS[st.key]||0) + " 条";
  stageArea.appendChild(span);
});

// Build tabs + panels
var nav = document.getElementById("tab-nav");
var panels = document.getElementById("panels");
TABS.forEach(function(tab, idx){
  var btn = document.createElement("button");
  btn.id = "btn-"+tab.id;
  btn.className = "tbtn" + (idx===0 ? " active" : "");
  btn.textContent = tab.label + " (" + tab.data.length + ")";
  btn.onclick = function(){ switchTab(tab.id); };
  nav.appendChild(btn);

  var panel = document.createElement("div");
  panel.id = "panel-"+tab.id;
  panel.className = "panel" + (idx===0 ? " active" : "");
  panel.innerHTML =
    '<div class="toolbar">'
    +'<input class="sinput" id="s-'+tab.id+'" type="text" placeholder="搜索 '+tab.label+'..." />'
    +'<span class="rc" id="c-'+tab.id+'"></span>'
    +'</div><div class="tw" id="t-'+tab.id+'"></div>';
  panels.appendChild(panel);

  document.getElementById("s-"+tab.id).addEventListener("input", function(){
    renderTable(tab.id, tab.cols, tab.data, tab.gradeKey, this.value);
  });
});

var sortState = {};

function renderTable(tabId, cols, data, gradeKey, filter){
  var filtered = filter ? data.filter(function(r){
    return Object.values(r).some(function(v){
      return String(v).toLowerCase().indexOf(filter.toLowerCase()) >= 0;
    });
  }) : data;

  var ss = sortState[tabId] || {};
  var sorted = filtered.slice();
  if(ss.col !== undefined){
    sorted.sort(function(a,b){
      var va = a[cols[ss.col]], vb = b[cols[ss.col]];
      var na = parseFloat(va), nb = parseFloat(vb);
      if(!isNaN(na) && !isNaN(nb)){ va=na; vb=nb; }
      else { va=String(va||""); vb=String(vb||""); }
      return ss.asc ? (va>vb?1:va<vb?-1:0) : (va<vb?1:va>vb?-1:0);
    });
  }

  var gi = cols.indexOf(gradeKey);
  var stageColIdx = cols.indexOf("评级阶段");
  var h = "<table><thead><tr>";
  for(var i=0; i<cols.length; i++){
    var cls = ss.col===i ? (ss.asc?"sa":"sd") : "";
    h += '<th class="'+cls+'" data-tab="'+tabId+'" data-col="'+i+'" onclick="handleSort(this)">'+cols[i]+'<span class="si"></span></th>';
  }
  h += "</tr></thead><tbody>";
  for(var j=0; j<sorted.length; j++){
    h += "<tr>";
    for(var i=0; i<cols.length; i++){
      var val = sorted[j][cols[i]];
      if(val===undefined||val===null) val="";
      if(i===gi && val){
        h += '<td class="gc-'+val+'"><span class="badge grade-'+val+'">'+val+'</span></td>';
      } else if(i===stageColIdx && val){
        var sc2 = "stage-"+val;
        h += '<td><span class="sbadge-cell '+sc2+'">'+String(val)+'</span></td>';
      } else {
        h += '<td title="'+String(val).replace(/"/g,"&quot;")+'">'+String(val)+'</td>';
      }
    }
    h += "</tr>";
  }
  h += "</tbody></table>";
  document.getElementById("t-"+tabId).innerHTML = h;
  document.getElementById("c-"+tabId).textContent = "共 " + sorted.length + " 条";
}

function handleSort(th){
  var tabId  = th.getAttribute("data-tab");
  var colIdx = parseInt(th.getAttribute("data-col"));
  var ss = sortState[tabId] || {};
  sortState[tabId] = {col: colIdx, asc: ss.col===colIdx ? !ss.asc : true};
  var tab = TABS.find(function(t){ return t.id===tabId; });
  var filter = document.getElementById("s-"+tabId).value;
  renderTable(tabId, tab.cols, tab.data, tab.gradeKey, filter);
}

function switchTab(id){
  TABS.forEach(function(t){
    document.getElementById("panel-"+t.id).classList.toggle("active", t.id===id);
    document.getElementById("btn-"+t.id).classList.toggle("active", t.id===id);
  });
}

// Initial render
TABS.forEach(function(tab){
  renderTable(tab.id, tab.cols, tab.data, tab.gradeKey, "");
});
"""

    html = (
        '<!DOCTYPE html>\n<html lang="zh-CN">\n<head>\n'
        '<meta charset="UTF-8">\n'
        '<meta name="viewport" content="width=device-width,initial-scale=1.0">\n'
        '<title>土耳其KOL达人库</title>\n'
        '<style>' + css + '</style>\n'
        '</head>\n<body>\n'
        '<header><h1>&#127481;&#127479; 土耳其KOL达人库</h1><p id="utime"></p></header>\n'
        '<div class="stats" id="stats-area"></div>\n'
        '<div class="grade-row" id="grade-area"></div>\n'
        '<div class="stage-row" id="stage-area"></div>\n'
        '<div class="tabs" id="tab-nav"></div>\n'
        '<div id="panels"></div>\n'
        '<script>\n' + data_block + js + '\n</script>\n'
        '</body>\n</html>\n'
    )

    with open(str(output_path), "w", encoding="utf-8") as fh:
        fh.write(html)
    print(f"[✓] HTML看板已生成: {output_path}")


# ─────────────────────────────────────────
# 加载/合并现有数据库
# ─────────────────────────────────────────
def load_db_detail():
    """加载现有数据库的视频明细"""
    if not DB_PATH.exists():
        return pd.DataFrame()
    try:
        df = pd.read_excel(DB_PATH, sheet_name="视频明细")
        return df.dropna(how="all")
    except Exception:
        return pd.DataFrame()


def load_db_pool():
    """加载现有数据库的达人池"""
    if not DB_PATH.exists():
        return pd.DataFrame()
    try:
        df = pd.read_excel(DB_PATH, sheet_name="达人池")
        return df.dropna(how="all")
    except Exception:
        return pd.DataFrame()


def ensure_review_columns(df):
    """向后兼容：确保新字段存在"""
    changed = False
    if "评级阶段" not in df.columns:
        df = df.copy()
        df["评级阶段"] = "7日终评"
        changed = True
    if "最后更新时间" not in df.columns:
        df = df.copy() if not changed else df
        df["最后更新时间"] = ""
    if "视频ID" not in df.columns:
        df["视频ID"] = ""
    return df


def dedup(existing_df, new_df):
    """去重追加：以(达人+宣发期+平台+发布日期)为唯一键"""
    if existing_df.empty:
        return new_df
    if new_df.empty:
        return existing_df

    key_cols = ["达人", "宣发期", "平台", "发布日期"]
    # 对API数据用视频ID去重
    if "视频ID" in existing_df.columns and "视频ID" in new_df.columns:
        existing_ids = set(existing_df["视频ID"].dropna().astype(str).unique()) - {"", "0"}
        new_df_filtered = new_df[
            ~new_df["视频ID"].astype(str).isin(existing_ids)
            | new_df["视频ID"].astype(str).isin(["", "0", "nan"])
        ]
        # 再用key_cols去重
        existing_keys = set(
            existing_df[key_cols].apply(lambda r: "|".join(r.astype(str)), axis=1)
        )
        new_df_filtered = new_df_filtered[
            ~new_df_filtered[key_cols].apply(
                lambda r: "|".join(r.astype(str)), axis=1
            ).isin(existing_keys)
        ]
        return pd.concat([existing_df, new_df_filtered], ignore_index=True)
    else:
        key_cols_avail = [c for c in key_cols if c in existing_df.columns and c in new_df.columns]
        existing_keys = set(
            existing_df[key_cols_avail].apply(lambda r: "|".join(r.astype(str)), axis=1)
        )
        new_df_filtered = new_df[
            ~new_df[key_cols_avail].apply(
                lambda r: "|".join(r.astype(str)), axis=1
            ).isin(existing_keys)
        ]
        return pd.concat([existing_df, new_df_filtered], ignore_index=True)


# ─────────────────────────────────────────
# 日期过滤
# ─────────────────────────────────────────
def filter_by_days(items, days):
    cutoff = datetime.utcnow() - timedelta(days=days)
    result = []
    for item in items:
        for v in item.get("videos", []):
            pd_str = v.get("publish_date") or v.get("create_time", "")
            try:
                dt = datetime.fromisoformat(pd_str.replace("Z", "+00:00"))
                if dt.replace(tzinfo=None) >= cutoff:
                    result.append(item)
                    break
            except Exception:
                pass
    return result


def filter_by_month(items, month_str):
    """month_str: 'YYYY-MM'"""
    result = []
    for item in items:
        for v in item.get("videos", []):
            pm = v.get("publish_month", "") or ""
            pd_str = v.get("publish_date", "") or ""
            if pm.startswith(month_str) or pd_str.startswith(month_str):
                result.append(item)
                break
    return result


# ─────────────────────────────────────────
# 主流程
# ─────────────────────────────────────────
def cmd_init():
    """从历史Excel初始化数据库"""
    print("[*] 初始化数据库...")
    print("  读取历史明细...")
    hist_df = load_history_detail()
    print(f"  历史明细: {len(hist_df)} 条")

    print("  读取4月数据...")
    apr_df = load_april_detail()
    print(f"  4月明细: {len(apr_df)} 条")

    detail_df = pd.concat([hist_df, apr_df], ignore_index=True)
    # 初始化时历史数据统一标记为7日终评（均已超过7天）
    detail_df["评级阶段"] = "7日终评"
    print(f"  合并总计: {len(detail_df)} 条（均标记为7日终评）")

    pool_df = load_pool()
    # 初始化时也自动更新达人池
    print("[*] 自动更新达人池...")
    pool_df = update_pool_from_detail(detail_df, pool_df)

    # 自动搜集粉丝量和联系方式
    print("\n[*] 自动搜集达人粉丝量和联系方式...")
    pool_df = collect_kol_info(pool_df)

    summary_df = compute_summary(detail_df)

    write_excel(detail_df, summary_df, pool_df, DB_PATH)
    print(f"[✓] 初始化完成！视频明细 {len(detail_df)} 条 | 达人汇总 {len(summary_df)} 人 | 达人池 {len(pool_df)} 人")


def cmd_update(days=None, month=None, fetch_all=False):
    """从API拉取新数据并更新数据库"""
    token = load_token()
    print("[*] 正在从Codeck API拉取TR达人数据...")
    all_items = fetch_all_tr_promotions(token)
    print(f"  API返回总记录: {len(all_items)}")

    # 过滤用于拉新视频的数据范围
    if days:
        items = filter_by_days(all_items, days)
        print(f"  过滤最近{days}天: {len(items)} 条")
    elif month:
        items = filter_by_month(all_items, month)
        print(f"  过滤{month}月份: {len(items)} 条")
    elif fetch_all:
        items = all_items
        print(f"  全量同步: {len(items)} 条")
    else:
        items = all_items

    # 新视频以「初评」阶段入库
    new_df = api_items_to_df(items, stage="初评")
    print(f"  展开视频明细: {len(new_df)} 条")

    # 加载现有数据库
    existing_df = load_db_detail()
    pool_df = load_db_pool()

    # 向后兼容：补全新字段
    if not existing_df.empty:
        existing_df = ensure_review_columns(existing_df)

    if existing_df.empty:
        print("[!] 数据库尚未初始化，请先运行 --init")
        merged_df = new_df if not new_df.empty else pd.DataFrame()
        added = len(merged_df)
    else:
        before = len(existing_df)
        merged_df = dedup(existing_df, new_df)
        added = len(merged_df) - before
        print(f"  新增视频: {added} 条（去重后）")

    if merged_df.empty:
        print("[!] 没有数据可处理")
        return

    # ── Bug 1 fix: fill empty publish dates from API / 宣发期 ──
    print("\n[*] 补填空发布日期...")
    merged_df = fill_missing_publish_dates(merged_df, all_items)

    # ── Bug 2 fix: correct stages that are inconsistent with publish date ──
    print("\n[*] 修正评级阶段...")
    merged_df = fix_stale_stages(merged_df)

    # ── 定期回查：3日 / 7日终评 ──
    print("\n[*] 执行3日/7日定期回查...")
    merged_df, updated_3day, updated_7day = run_periodic_reviews(merged_df, all_items)
    print(f"  3日回查候选: {len(updated_3day)} 条  |  7日终评候选: {len(updated_7day)} 条")

    summary_df = compute_summary(merged_df)

    # ── 自动更新达人池 ──
    print("\n[*] 自动更新达人池...")
    pool_df = update_pool_from_detail(merged_df, pool_df)

    # ── 自动搜集粉丝量和联系方式（仅对空缺项）──
    print("\n[*] 自动搜集达人粉丝量和联系方式...")
    pool_df = collect_kol_info(pool_df)

    write_excel(merged_df, summary_df, pool_df, DB_PATH)

    # ── 输出更新报告 ──
    print("\n===== 更新报告 =====")
    print(f"新增视频: {added} 条")

    print(f"\n3日回查更新: {len(updated_3day)} 条")
    for rec in updated_3day:
        change = f"[评级: {rec['旧评级']} → {rec['新评级']}]" if rec["旧评级"] != rec["新评级"] else "[无变化]"
        src = "(API已更新)" if rec["API更新"] else "(无API数据，仅推进阶段)"
        print(f"  {rec['达人']}  {change}  {src}")

    print(f"\n7日终评更新: {len(updated_7day)} 条")
    for rec in updated_7day:
        change = f"[评级: {rec['旧评级']} → {rec['新评级']}]" if rec["旧评级"] != rec["新评级"] else "[无变化]"
        src = "(API已更新)" if rec["API更新"] else "(无API数据，仅推进阶段)"
        print(f"  {rec['达人']}  {change}  {src}")

    if not new_df.empty:
        grade_counts = new_df["评级"].value_counts().to_dict()
        print(f"\n新拉取数据评级分布: {grade_counts}")
    print("====================")


def main():
    parser = argparse.ArgumentParser(description="土耳其KOL达人库更新工具")
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--init", action="store_true", help="从历史Excel初始化数据库")
    group.add_argument("--days", type=int, help="拉取最近N天数据")
    group.add_argument("--month", type=str, help="拉取指定月份数据 (格式: YYYY-MM)")
    group.add_argument("--all", action="store_true", dest="fetch_all", help="拉取全部TR数据")
    args = parser.parse_args()

    if args.init:
        cmd_init()
    elif args.days:
        cmd_update(days=args.days)
    elif args.month:
        cmd_update(month=args.month)
    elif args.fetch_all:
        cmd_update(fetch_all=True)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
